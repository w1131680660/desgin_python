import datetime
import json
import re
from operator import itemgetter

import requests
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from settings import conf_fun


def create_order(container_num, factory, country, site, user_name):
    print("########################## 生成下单表")
    print("准备生成文件")
    # 当前的日期
    now_date = datetime.datetime.strftime(datetime.datetime.now() + datetime.timedelta(hours=8), "%Y-%m-%d")
    # operating_data 对于货柜号和工厂国家的数据
    sql = " SELECT * FROM operating_data WHERE container_num='{0}' and factory ='{1}' and country ='{2}' and site ='{3}'" \
        .format(container_num, factory, country, site)
    print(sql)
    sc_sql = " SELECT * from schedule_container WHERE container_num='{0}' and country ='{1}' and site ='{2}'" \
        .format(container_num, country, site)

    supply_sql = " SELECT *  FROM supply_chain_data WHERE supply_chain_name='{0}'".format(factory)
    order_sql = " SELECT * FROM order_distribution WHERE container_code ='{0}' AND factory ='{1}'".format(container_num,
                                                                                                          factory)
    select_distribution_res_two = conf_fun.connect_mysql_supply(order_sql, type='dict')

    re_sc_data = conf_fun.connect_mysql_operation(sc_sql, type='dict')
    re_operating_data = conf_fun.connect_mysql_operation(sql, type='dict')
    re_supply_data = conf_fun.connect_mysql_supply(supply_sql, type='dict')
    re_operating_dict = re_operating_data[0]
    print(re_operating_dict)
    id = re_operating_dict.get('id')
    user_name_1 = re_operating_dict.get('user_name') if re_operating_dict.get('user_name') else user_name
    schedule_date = re_sc_data[0].get('schedule_date')

    iphone_sql = " SELECT * FROM userinfo WHERE real_name ='{0}'".format(user_name.strip())
    re_iphone = conf_fun.connect_mysql_operation(iphone_sql, type='dict')
    iphone = re_iphone[0].get('phone') if re_iphone else ''
    supply_chain_allname = re_supply_data[0].get('supply_chain_allname') if re_supply_data else ''
    supply_chain_site = re_supply_data[0].get('supply_chain_site') if re_supply_data else ''

    target_number = re_operating_dict.get('calculation_file_path').split('/')[-1].split('-')[1]

    select_country_sql = "select * from country_contrast where country_cn='%s';" % country
    select_site_sql = "select * from area_contrast where area_cn='%s';" % (site)
    select_country_res = conf_fun.connect_mysql_operation(select_country_sql, dbs="operation", type='dict')
    select_site_res = conf_fun.connect_mysql_operation(select_site_sql, dbs="operation", type="dict")

    try:
        en_country = select_country_res[0]['country']
        en_site = select_site_res[0]['area']
    except Exception as e:
        print("错误: ", "国家站点对照表信息不全!")
        res = {"code": 4041, "msg": "国家站点对照表信息不全!请联系IT部处理!"}
        return res
    # 生成文件
    wb = Workbook()
    ws = wb.active
    ws.title = '订单模型'

    ws.merge_cells('A1:I1')
    ws["A1"] = "工厂下单确认书"

    ws.merge_cells('A2:I2')
    ws["A2"] = "一、订单基本信息"

    ws.merge_cells('A3:B3')
    ws["A3"] = "货运编号"
    ws.merge_cells('C3:D3')
    # 货柜号-测算表中间那串数字站点（英文）国家（英文）
    ws["C3"] = container_num + "-" + target_number + en_site + en_country

    ws["E3"] = "下单人员"
    # operating_data里的user_name
    ws["F3"] = user_name_1
    ws["G3"] = "下单人员电话"
    ws.merge_cells('H3:I3')
    # userinfo里的phone
    ws["H3"] = iphone

    ws.merge_cells('A4:B4')
    ws["A4"] = "下单日期"
    ws.merge_cells('C4:D4')
    # 现在的日期
    ws["C4"] = now_date
    ws["E4"] = "质检日期"
    ws["G4"] = "装柜日期"
    ws.merge_cells('H4:I4')
    # order_distribution里的dates
    ws["H4"] = schedule_date

    ws.merge_cells('A5:B5')
    ws["A5"] = "供应商名称"
    ws.merge_cells('C5:D5')
    # supply_chain_data里的supply_chain_allname
    ws["C5"] = supply_chain_allname
    ws["E5"] = "工厂跟单"
    ws["G5"] = "跟单电话"
    ws.merge_cells('H5:I5')

    ws.merge_cells('A6:B6')
    ws["A6"] = "供应商地址"
    ws.merge_cells('C6:I6')
    # supply_chain_data里的supply_chain_site
    ws["C6"] = supply_chain_site

    ws.merge_cells('A7:B7')
    ws["A7"] = "本公司名称"
    ws.merge_cells('C7:I7')
    ws["C7"] = "杭州中睿实业有限公司"

    ws.merge_cells('A8:B8')
    ws["A8"] = "办公地址"
    ws.merge_cells('C8:I8')

    ws.merge_cells('A9:I9')
    ws["A9"] = "二、订单备注"

    ws["A10"] = "质检要求："
    ws["A11"] = "1"
    ws["A12"] = "2"
    ws["A13"] = "3"
    ws["A14"] = "4"
    ws["A15"] = "5"
    ws["A16"] = "6"
    ws["A17"] = "违约条款"
    ws["A18"] = "1"
    ws["A19"] = "2"
    ws["A20"] = "付款条款"
    ws["A21"] = "1"

    ws.merge_cells('B11:I11')
    ws["B11"] = "明确样品总数以及抽检比例（5%）左右即抽检样本数=样品总数*5%"
    ws.merge_cells('B12:I12')
    ws["B12"] = "明确检测项目以及合格标准，提前准备好相应的工具确保这些工具精确、可用。并且明确这些工具的使用方法。"
    ws.merge_cells('B13:I13')
    ws["B13"] = "合理划分取样区域，抽检的样品要具有代表性，习惯按款划分：每款抽检对应样本数=抽检样本数/款式总数，再在这个款的样品中随机抽检对应样本数。"
    ws.merge_cells('B14:I14')
    ws["B14"] = "如抽检过程中有1-2个不合格，再随机抽查10个，这10个样品中大于等于9个合格即完成质量确认，否则拒收。并且及时记下不合格产品数。"
    ws.merge_cells('B15:I15')
    ws["B15"] = "计算不合格产品数占抽检样品数的比例：不合格产品数/抽检样品总数*100%"
    ws.merge_cells('B16:I16')
    ws["B16"] = "必要时带回代表样本（比如衣柜：三种样式的门板，pp塑料连接器，磁扣，小木锤，外包装纸箱等）。"
    ws.merge_cells('B18:I18')
    ws["B18"] = "如延迟交付3天及以上，违约金为剩余货款的1%每天。"
    ws.merge_cells('B19:I19')
    ws["B19"] = "如出现供货细节与订单不符，并已盖章出货，供应商需承担赔偿责任。"
    ws.merge_cells('B21:I21')
    ws["B21"] = "以双方最近一次实际约定为准"

    ws["A22"] = "序号"
    ws["B22"] = "sku"
    ws["C22"] = "货名"
    ws["D22"] = "规格说明"
    ws["E22"] = "配件明细"
    ws["F22"] = "纸箱尺寸"
    ws["G22"] = "单价"
    ws["H22"] = "下单数量"
    ws["I22"] = "总价"

    primary_title_font = Font(name="宋体", size=26, bold=True)
    secondary_title_font = Font(name="宋体", size=18)
    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    fill = PatternFill(fill_type="solid", fgColor="A6A6A6")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws.cell(row=1, column=1).font = primary_title_font
    ws.cell(row=1, column=1).alignment = alignment

    for i in range(3, 11):
        ws.cell(row=i, column=1).fill = fill

    ws.cell(row=2, column=1).font = secondary_title_font
    ws.cell(row=2, column=1).fill = fill

    ws.cell(row=9, column=1).font = secondary_title_font

    for i in range(3, 6):
        ws.cell(row=i, column=5).fill = fill

    for i in range(3, 6):
        ws.cell(row=i, column=7).fill = fill

    ws.cell(row=17, column=1).fill = fill
    ws.cell(row=20, column=1).fill = fill

    for i in range(1, 10):
        ws.cell(row=22, column=i).fill = fill

    # 循环写入数据
    all_num = 0
    re_list = []
    pattern = '\d+'
    for i in select_distribution_res_two:
        product_name = i.get('product_name')
        if product_name:
            re_name = re.findall(pattern, product_name)
            re_name = int(re_name[0]) if re_name else 0
            i['re_name'] = re_name
        product_number = i.get('product_number')
        if product_number:
            product_number_list = product_number.split('-')[0:-1]
            product_number_re = '-'.join(product_number_list)
            i['product_number_re'] = product_number_re
        re_list.append(i)
    select_distribution_res_two = sorted(re_list, key=itemgetter('product_number_re', 're_name'))
    if not select_distribution_res_two:
        return {'code': 206, 'msg': '货柜号对应的工厂还未分配'}
    for i in range(len(select_distribution_res_two)):
        select_part_sql = "select DISTINCT part_name,part_number,part_code  from product_parts " \
                          "where product_code='%s' ORDER BY part_code ;" \
                          % select_distribution_res_two[i]['product_number']
        print(select_part_sql)
        select_part_res = conf_fun.connect_mysql_supply(select_part_sql, dbs="supply_chain", type="dict")

        part_name = ''
        for j in select_part_res:
            part_name += "%s   %s   \n" % (j.get('part_name'), j.get('part_number'))

        # 查询product_message
        select_package_sql = "select * from product_message where product_code='%s';" \
                             % select_distribution_res_two[i]['product_number']
        select_package_res = conf_fun.connect_mysql_operation(select_package_sql, "operation", "dict")
        product_package_size = select_package_res[0]['product_package_size']

        ws["A" + str(i + 23)] = i + 1
        ws["A" + str(i + 23)].alignment = alignment
        ws.cell(row=i + 23, column=1).fill = fill
        ws["B" + str(i + 23)] = select_distribution_res_two[i]['sku']
        ws["B" + str(i + 23)].alignment = alignment
        ws["C" + str(i + 23)] = select_distribution_res_two[i]['product_name']
        ws["C" + str(i + 23)].alignment = alignment
        ws["D" + str(i + 23)] = '加深/32丝'
        ws["D" + str(i + 23)].alignment = alignment
        # 配件明细
        ws["E" + str(i + 23)] = part_name
        ws["E" + str(i + 23)].alignment = alignment
        # 纸箱尺寸
        ws["F" + str(i + 23)] = product_package_size
        ws["F" + str(i + 23)].alignment = alignment
        #                        # 单价
        ws["G" + str(i + 23)] = ''
        ws["G" + str(i + 23)].alignment = alignment
        # 下单数量
        ws["H" + str(i + 23)] = select_distribution_res_two[i]['distribution_num']
        ws["H" + str(i + 23)].alignment = alignment

        all_num += float(select_distribution_res_two[i]['distribution_num'])

    ws["A" + str(len(select_distribution_res_two) + 24)] = "合计"
    ws["H" + str(len(select_distribution_res_two) + 24)] = all_num

    for i in range(1, 10):
        ws.cell(row=len(select_distribution_res_two) + 24, column=i).fill = fill

    for i in range(1, 10):
        for j in range(1, len(select_distribution_res_two) + 25):
            ws.cell(row=j, column=i).border = thin_border

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12

    # 保存文件
    local_path = '/home/beyong_supply_chain/static/data/place_order/'
    import os
    if not os.path.exists(local_path):
        print(123)
        os.makedirs(local_path)
    filename = site + country + "-" + factory + '-' + now_date[5:0].replace(
        "-", "") + '-' + container_num + "-下单表.xlsm"
    print("保存文件名: ", filename)
    wb.save(local_path + filename)

    url = 'https://www.beyoung.group/file_upload/'
    remote_path = 'operation/operating_data/place_order/'
    data = {'path': remote_path}
    res = requests.post(url, data, files={'file': open(local_path + filename, 'rb')})
    res_data = json.loads(res.text)
    if res_data['code'] != 200:
        res = {"code": 4041, "msg": "上传下单表失败!联系IT部!"}
        return res

    else:
        remote_path += '%s' % filename
        update_sql = " UPDATE operating_data SET place_order='{0}' ,place_order_date ='{1}' " \
                     " WHERE id='{2}'".format(remote_path, now_date, id)
        print(update_sql)
        conf_fun.connect_mysql_operation(update_sql)
        return {'code': 200}


def order_from(request):
    print(request)
    data = request.GET
    container_num = data.get('container_num')
    factory = data.get('factory')
    country = data.get('country')
    site = data.get('site')
    # user_name =unquote(request.META.get('HTTP_AUTHORIZATION')).split('@')[0]
    user_name = '黄'
    print(container_num, factory, country, site)
    judge_sql = " SELECT DISTINCT product_type FROM order_distribution " \
                "WHERE container_code ='{0}' AND factory ='{1}' And country='2' AND store ='{3}';" \
        .format(container_num, factory, country, site)
    re_judge_data = conf_fun.connect_mysql_supply(judge_sql, type='dict')
    for i in re_judge_data:
        if i.get('product_type') != '魔片':
            ret = {'code': 206, 'msg': '非魔片不会生成下单表'}
            return JsonResponse(ret)
    try:
        ret = create_order(container_num, factory, country, site, user_name)
    except:
        ret = {'code': 500, 'msg': '下单表生成对应的商品信息有误，请检查对应工厂的下单信息'}
    return JsonResponse(ret)
