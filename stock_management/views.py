from django.shortcuts import render
from urllib.parse import unquote
from django.http import JsonResponse
import os
import time
import pymysql
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
# Create your views here.


def demo(request):
    return render(request,"home.html")


def is_open(request):
    area = unquote(request.META.get('HTTP_AUTHORIZATION')).split('@')[1]
    if 'all' == area:
        return JsonResponse({"code": 200})
    else:
        return JsonResponse({"code": 403})


def connect_mysql(sql_text, dbs='operation', type='tuple'):
    conn = pymysql.Connect(host='gz-cdb-lwqgjirt.sql.tencentcdb.com', port=59656, user='beyoungsql', passwd='Bymy2021_',db=dbs)
    if type == 'tuple':
        cursor = conn.cursor()
    else:
        cursor = conn.cursor(pymysql.cursors.DictCursor)
    cursor.execute(sql_text)
    response = cursor.fetchall()
    conn.commit()
    cursor.close()
    conn.close()
    return response


def master_upload_file(files, file_path):
    url = 'https://www.beyoung.group/file_upload/'
    path2 = os.path.join(r'operation/operating_data/', file_path)
    data = {'path': path2}
    print(data)
    res = requests.post(url, data, files={'file': files})
    path3 = os.path.join(r'operation/operating_data/', file_path, str(files))
    print('这是什么路径\n', path3)
    return path3


def auto_place_order(request):
    container_num = request.GET.get('container_num') # D577-2156
    user_name = unquote(request.META.get('HTTP_AUTHORIZATION')).split('@')[0]
    place_order_time = request.GET.get('place_order_time')
    loading_time = request.GET.get('loading_time')
    area = request.GET.get('area')
    cabinet_type = request.GET.get('cabinet_type')
    # container_num = 'D577-2156'
    # user_name = '罗秋丽'
    # place_order_time = '20210112'
    # loading_time = '202101'
    # area = '中睿德国'

    times = str(time.strftime("%Y-%m%d", time.localtime())).split('-')[1]
    times1 = str(time.strftime("%Y-%m-%d", time.localtime()))

    container = container_num.split('-')[0]

    if '中睿' in area:
        site = 'ZR'
    elif '爱瑙' in area:
        site = 'AN'
    elif '胤佑' in area:
        site = 'YY'
    elif '京汇' in area:
        site = 'JH'
    elif '利百锐' in area:
        site = 'LBR'

    if '美国' in area:
        country = 'US'
    elif '英国' in area:
        country = 'UK'
    elif '德国' in area:
        country = 'DE'
    elif '欧洲' in area:
        country = 'EU'
    elif '加拿大' in area:
        country = 'CA'
    elif '日本' in area:
        country = 'JP'
    elif '澳洲' in area:
        country = 'AU'
    elif '墨西哥' in area:
        country = 'MX'

    sql = "select sum(surplus_distribution_order),container_code from order_integrate  group by container_code having " \
          "container_code='{}'"
    sql = sql.format(container)
    print(sql)
    res = connect_mysql(sql, dbs='supply_chain')
    if res[0][0] > 0:
        return JsonResponse({"code": 400, "msg": "供应链"})

    sql = "select factory from order_distribution where container_code='{}'"
    sql = sql.format(container)
    factory_res = connect_mysql(sql, dbs='supply_chain')
    if len(factory_res[0]) > 1:
        place_order_all = ''
        for factory_list in factory_res:
            sql = "select sku,distribution_num from order_distribution where container_code='{}' and factory='{}'"
            sql = sql.format(factory_list[0], container)
            sku_data = connect_mysql(sql, dbs='supply_chain')
            if len(sku_data) == 0:
                return JsonResponse({"code": 400, "msg": "没有该货柜信息！"})

            sql = "select phone from userinfo where real_name='{}'"
            sql = sql.format(user_name)
            user_tel = connect_mysql(sql)

            sql = "select supply_chain_allname,supply_chain_site from supply_chain_data where supply_chain_name='{}'"
            sql = sql.format(factory_list[0])
            supply_chain_data = connect_mysql(sql, dbs='supply_chain')

            wb = Workbook()
            ws = wb.active
            ws.title = '订单模型'

            ws.merge_cells('A1:I1')
            ws["A1"] = "工厂下单确认书"

            ws.merge_cells('A2:I2')
            ws["A2"] = "一、订单基本信息"

            ws.merge_cells('A3:B3')
            ws["A3"] = "货运编号"
            ws.merge_cells('C3:D3')
            ws["C3"] = container_num + site + country
            ws["E3"] = "运营人员"
            ws["F3"] = user_name
            ws["G3"] = "运营电话"
            ws.merge_cells('H3:I3')
            ws["H3"] = user_tel[0][0]

            ws.merge_cells('A4:B4')
            ws["A4"] = "下单日期"
            ws.merge_cells('C4:D4')
            ws["C4"] = place_order_time
            ws["E4"] = "质检日期"
            ws["G4"] = "装柜日期"
            ws.merge_cells('H4:I4')
            ws["H4"] = loading_time

            ws.merge_cells('A5:B5')
            ws["A5"] = "供应商名称"
            ws.merge_cells('C5:D5')
            ws["C5"] = supply_chain_data[0][0]
            ws["E5"] = "工厂跟单"
            ws["G5"] = "跟单电话"
            ws.merge_cells('H5:I5')

            ws.merge_cells('A6:B6')
            ws["A6"] = "供应商地址"
            ws.merge_cells('C6:I6')
            ws["C6"] = supply_chain_data[0][1]

            ws.merge_cells('A7:B7')
            ws["A7"] = "本公司名称"
            ws.merge_cells('C7:I7')
            ws["C7"] = "杭州中睿实业有限公司"

            ws.merge_cells('A8:B8')
            ws["A8"] = "办公地址"
            ws.merge_cells('C8:I8')

            ws.merge_cells('A9:I9')
            ws["A9"] = "二、订单备注"

            ws["A10"] = "质检要求："
            ws["A11"] = "1"
            ws["A12"] = "2"
            ws["A13"] = "3"
            ws["A14"] = "4"
            ws["A15"] = "5"
            ws["A16"] = "6"
            ws["A17"] = "违约条款"
            ws["A18"] = "1"
            ws["A19"] = "2"
            ws["A20"] = "付款条款"
            ws["A21"] = "1"

            ws.merge_cells('B11:I11')
            ws["B11"] = "明确样品总数以及抽检比例（5%）左右即抽检样本数=样品总数*5%"
            ws.merge_cells('B12:I12')
            ws["B12"] = "明确检测项目以及合格标准，提前准备好相应的工具确保这些工具精确、可用。并且明确这些工具的使用方法。"
            ws.merge_cells('B13:I13')
            ws["B13"] = "合理划分取样区域，抽检的样品要具有代表性，习惯按款划分：每款抽检对应样本数=抽检样本数/款式总数，再在这个款的样品中随机抽检对应样本数。"
            ws.merge_cells('B14:I14')
            ws["B14"] = "如抽检过程中有1-2个不合格，再随机抽查10个，这10个样品中大于等于9个合格即完成质量确认，否则拒收。并且及时记下不合格产品数。"
            ws.merge_cells('B15:I15')
            ws["B15"] = "计算不合格产品数占抽检样品数的比例：不合格产品数/抽检样品总数*100%"
            ws.merge_cells('B16:I16')
            ws["B16"] = "必要时带回代表样本（比如衣柜：三种样式的门板，pp塑料连接器，磁扣，小木锤，外包装纸箱等）。"
            ws.merge_cells('B18:I18')
            ws["B18"] = "如延迟交付3天及以上，违约金为剩余货款的1%每天。"
            ws.merge_cells('B19:I19')
            ws["B19"] = "如出现供货细节与订单不符，并已盖章出货，供应商需承担赔偿责任。"
            ws.merge_cells('B21:I21')
            ws["B21"] = "以双方最近一次实际约定为准"

            ws["A22"] = "序号"
            ws["B22"] = "sku"
            ws["C22"] = "货名"
            ws["D22"] = "规格说明"
            ws["E22"] = "配件明细"
            ws["F22"] = "纸箱尺寸"
            ws["G22"] = "单价"
            ws["H22"] = "下单数量"
            ws["I22"] = "总价"

            print('test.xlsx')

            font = Font(name="宋体", size=18)
            font1 = Font(name="宋体", size=26, bold=True)
            alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            fill = PatternFill(fill_type="solid", fgColor="A6A6A6")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            ws.cell(row=1, column=1).font = font1
            ws.cell(row=1, column=1).alignment = alignment

            for i in range(3, 11):
                ws.cell(row=i, column=1).fill = fill

            ws.cell(row=2, column=1).font = font
            ws.cell(row=2, column=1).fill = fill

            ws.cell(row=9, column=1).font = font

            for i in range(3, 6):
                ws.cell(row=i, column=5).fill = fill

            for i in range(3, 6):
                ws.cell(row=i, column=7).fill = fill


            ws.cell(row=17, column=1).fill = fill
            ws.cell(row=20, column=1).fill = fill

            for i in range(1, 10):
                ws.cell(row=22, column=i).fill = fill

            all_num = 0
            for i in range(len(sku_data)):
                sql = "select commodity_name,product_package_size from commodity_information ci join product_message pm on " \
                      "ci.product_code=pm.product_code where sku='{}'"
                sql = sql.format(sku_data[i][0])
                sku_res = connect_mysql(sql)
                part_name = ''
                sql = "select part_name from commodity_information ci join product_components pc on ci.product_code=" \
                      "pc.product_code where sku='{}'"
                sql = sql.format(sku_data[i][0])
                part_res = connect_mysql(sql)
                for j in part_res:
                    part_name += j[0] + '\n'
                ws["A" + str(i + 23)] = i + 1
                ws.cell(row=i + 23, column=1).fill = fill
                ws["B" + str(i + 23)] = sku_data[i][0]
                ws["C" + str(i + 23)] = sku_res[0][0]
                ws["D" + str(i + 23)] = '加深/32丝'
                ws["E" + str(i + 23)] = part_name
                ws["F" + str(i + 23)] = sku_res[0][1]
                ws["H" + str(i + 23)] = sku_data[i][1]
                all_num += int(sku_data[i][1])

            ws["A" + str(len(sku_data) + 24)] = "合计"
            ws["H" + str(len(sku_data) + 24)] = all_num


            for i in range(1, 10):
                ws.cell(row=len(sku_data) + 24, column=i).fill = fill

            for i in range(1, 10):
                for j in range(1, len(sku_data) + 25):
                    ws.cell(row=j, column=i).border = thin_border

            for _row in ws.range('A1:L' + str(ws.get_highest_row())):
                for _cell in _row:
                    _cell.style.borders.left.border_style = Border.BORDER_THIN
                    _cell.style.borders.right.border_style = Border.BORDER_THIN
                    _cell.style.borders.top.border_style = Border.BORDER_THIN
                    _cell.style.borders.bottom.border_style = Border.BORDER_THIN
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 12

            path = '/home/by_operate/static/data/place_order/'
            filename = area + '-' + container_num.split('-')[1] + factory_list[0] + '-' + user_name + '-' + times + '-' + \
                       container_num.split('-')[0] + '-' + cabinet_type + '-下单表.xlsm'
            wb.save(path + filename)

            with open(path + filename, 'rb') as f:
                path1 = master_upload_file(f, 'place_order')
                place_order_all += path1 + '@'

        sql = "update operating_data set place_order='{}',place_order_date='{}',place_order_state='已通过' where " \
              "container_num = '{}'"
        sql = sql.format(place_order_all, times1, container_num.split('-')[0])
        connect_mysql(sql)

    else:
        sql = "select sku,distribution_num from order_distribution where container_code='{}' and factory='{}'"
        sql = sql.format(factory_res[0][0], container)
        sku_data = connect_mysql(sql, dbs='supply_chain')

        sql = "select phone from userinfo where real_name='{}'"
        sql = sql.format(user_name)
        user_tel = connect_mysql(sql)

        sql = "select supply_chain_allname,supply_chain_site from supply_chain_data where supply_chain_name='{}'"
        sql = sql.format(factory_res[0][0])
        supply_chain_data = connect_mysql(sql, dbs='supply_chain')

        wb = Workbook()
        ws = wb.active
        ws.title = '订单模型'

        ws.merge_cells('A1:I1')
        ws["A1"] = "工厂下单确认书"

        ws.merge_cells('A2:I2')
        ws["A2"] = "一、订单基本信息"

        ws.merge_cells('A3:B3')
        ws["A3"] = "货运编号"
        ws.merge_cells('C3:D3')
        ws["C3"] = container_num + site + country
        ws["E3"] = "运营人员"
        ws["F3"] = user_name
        ws["G3"] = "运营电话"
        ws.merge_cells('H3:I3')
        ws["H3"] = user_tel[0][0]

        ws.merge_cells('A4:B4')
        ws["A4"] = "下单日期"
        ws.merge_cells('C4:D4')
        ws["C4"] = place_order_time
        ws["E4"] = "质检日期"
        ws["G4"] = "装柜日期"
        ws.merge_cells('H4:I4')
        ws["H4"] = loading_time

        ws.merge_cells('A5:B5')
        ws["A5"] = "供应商名称"
        ws.merge_cells('C5:D5')
        ws["C5"] = supply_chain_data[0][0]
        ws["E5"] = "工厂跟单"
        ws["G5"] = "跟单电话"
        ws.merge_cells('H5:I5')

        ws.merge_cells('A6:B6')
        ws["A6"] = "供应商地址"
        ws.merge_cells('C6:I6')
        ws["C6"] = supply_chain_data[0][1]

        ws.merge_cells('A7:B7')
        ws["A7"] = "本公司名称"
        ws.merge_cells('C7:I7')
        ws["C7"] = "杭州中睿实业有限公司"

        ws.merge_cells('A8:B8')
        ws["A8"] = "办公地址"
        ws.merge_cells('C8:I8')

        ws.merge_cells('A9:I9')
        ws["A9"] = "二、订单备注"

        ws.merge_cells('B10:I10')
        ws.merge_cells('B17:I17')
        ws.merge_cells('B20:I20')
        ws["A10"] = "质检要求："
        ws["A11"] = "1"
        ws["A12"] = "2"
        ws["A13"] = "3"
        ws["A14"] = "4"
        ws["A15"] = "5"
        ws["A16"] = "6"
        ws["A17"] = "违约条款"
        ws["A18"] = "1"
        ws["A19"] = "2"
        ws["A20"] = "付款条款"
        ws["A21"] = "1"

        ws.merge_cells('B11:I11')
        ws["B11"] = "明确样品总数以及抽检比例（5%）左右即抽检样本数=样品总数*5%"
        ws.merge_cells('B12:I12')
        ws["B12"] = "明确检测项目以及合格标准，提前准备好相应的工具确保这些工具精确、可用。并且明确这些工具的使用方法。"
        ws.merge_cells('B13:I13')
        ws["B13"] = "合理划分取样区域，抽检的样品要具有代表性，习惯按款划分：每款抽检对应样本数=抽检样本数/款式总数，再在这个款的样品中随机抽检对应样本数。"
        ws.merge_cells('B14:I14')
        ws["B14"] = "如抽检过程中有1-2个不合格，再随机抽查10个，这10个样品中大于等于9个合格即完成质量确认，否则拒收。并且及时记下不合格产品数。"
        ws.merge_cells('B15:I15')
        ws["B15"] = "计算不合格产品数占抽检样品数的比例：不合格产品数/抽检样品总数*100%"
        ws.merge_cells('B16:I16')
        ws["B16"] = "必要时带回代表样本（比如衣柜：三种样式的门板，pp塑料连接器，磁扣，小木锤，外包装纸箱等）。"
        ws.merge_cells('B18:I18')
        ws["B18"] = "如延迟交付3天及以上，违约金为剩余货款的1%每天。"
        ws.merge_cells('B19:I19')
        ws["B19"] = "如出现供货细节与订单不符，并已盖章出货，供应商需承担赔偿责任。"
        ws.merge_cells('B21:I21')
        ws["B21"] = "以双方最近一次实际约定为准"

        ws["A22"] = "序号"
        ws["B22"] = "sku"
        ws["C22"] = "货名"
        ws["D22"] = "规格说明"
        ws["E22"] = "配件明细"
        ws["F22"] = "纸箱尺寸"
        ws["G22"] = "单价"
        ws["H22"] = "下单数量"
        ws["I22"] = "总价"

        font = Font(name="宋体", size=18)
        font1 = Font(name="宋体", size=26, bold=True)
        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        fill = PatternFill(fill_type="solid", fgColor="A6A6A6")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ws.cell(row=1, column=1).font = font1
        ws.cell(row=1, column=1).alignment = alignment

        for i in range(3, 11):
            ws.cell(row=i, column=1).fill = fill

        ws.cell(row=2, column=1).font = font
        ws.cell(row=2, column=1).fill = fill

        ws.cell(row=9, column=1).font = font

        for i in range(11, 17):
            ws.cell(row=i, column=1).alignment = alignment

        ws.cell(row=18, column=1).alignment = alignment
        ws.cell(row=19, column=1).alignment = alignment
        ws.cell(row=21, column=1).alignment = alignment

        for i in range(3, 6):
            ws.cell(row=i, column=5).fill = fill

        for i in range(3, 6):
            ws.cell(row=i, column=7).fill = fill

        ws.cell(row=17, column=1).fill = fill
        ws.cell(row=20, column=1).fill = fill

        for i in range(1, 10):
            ws.cell(row=22, column=i).fill = fill

        all_num = 0
        for i in range(len(sku_data)):
            sql = "select commodity_name,product_package_size from commodity_information ci join product_message pm on " \
                  "ci.product_code=pm.product_code where sku='{}'"
            sql = sql.format(sku_data[i][0])
            sku_res = connect_mysql(sql)
            part_name = ''
            sql = "select part_name from commodity_information ci join product_components pc on ci.product_code=" \
                  "pc.product_code where sku='{}'"
            sql = sql.format(sku_data[i][0])
            part_res = connect_mysql(sql)
            for j in part_res:
                part_name += j[0] + '\n'
            ws["A" + str(i + 23)] = i + 1
            ws.cell(row=i + 23, column=1).fill = fill
            ws["B" + str(i + 23)] = sku_data[i][0]
            ws["C" + str(i + 23)] = sku_res[0][0]
            ws["D" + str(i + 23)] = '加深/32丝'
            ws["E" + str(i + 23)] = part_name
            ws["F" + str(i + 23)] = sku_res[0][1]
            ws["H" + str(i + 23)] = sku_data[i][1]
            all_num += int(sku_data[i][1])

        ws["A" + str(len(sku_data) + 24)] = "合计"
        ws["H" + str(len(sku_data) + 24)] = all_num

        for i in range(1, 10):
            ws.cell(row=len(sku_data) + 24, column=i).fill = fill

        for i in range(1, 10):
            for j in range(1, len(sku_data) + 25):
                ws.cell(row=j, column=i).border = thin_border

        for _row in ws.range('A1:L' + str(ws.get_highest_row())):
            for _cell in _row:
                _cell.style.borders.left.border_style = Border.BORDER_THIN
                _cell.style.borders.right.border_style = Border.BORDER_THIN
                _cell.style.borders.top.border_style = Border.BORDER_THIN
                _cell.style.borders.bottom.border_style = Border.BORDER_THIN

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12

        path = '/home/by_operate/static/data/place_order/'
        filename = area + '-' + container_num.split('-')[1] + factory_res[0][0] + '-' + user_name + '-' + times + '-' + \
                   container_num.split('-')[0] + '-' + cabinet_type + '-下单表.xlsm'

        wb.save(path + filename)

        with open(path + filename, 'rb') as f:
            path1 = master_upload_file(f, 'place_order')
        sql = "update operating_data set place_order='{}',place_order_date='{}',place_order_state='已通过' where " \
              "container_num = '{}'"
        sql = sql.format(path1, times1, container_num.split('-')[0])
        connect_mysql(sql)

    return JsonResponse({'code': 200, 'msg': '生成成功！'})