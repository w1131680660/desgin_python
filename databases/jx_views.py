import re

import chardet
from django.http import JsonResponse, FileResponse
import pymysql
import os
import requests
import datetime
import time
import openpyxl
import json
import pandas as pd
from settings import conf_fun

# # 连接运营系统数据库
# def connect_mysql10(sql_text, dbs='operation', type='tuple'):
#     conn = pymysql.Connect(host='106.52.43.196', port=3306, user='beyoungsql', passwd='Bymy2021.', db=dbs)
#     if type == 'tuple':
#         cursor = conn.cursor()
#     else:
#         cursor = conn.cursor(pymysql.cursors.DictCursor)
#     cursor.execute(sql_text)
#     response = cursor.fetchall()
#     conn.commit()
#     cursor.close()
#     conn.close()
#     return response
#
#
# # 连接主系统数据库
# def connect_mysql1(sql_text, dbs='reports', type='tuple'):
#     conn = pymysql.Connect(host='106.53.250.215', port=3306, user='beyoungsql', passwd='Bymy2021.', db=dbs)
#     if type == 'tuple':
#         cursor = conn.cursor()
#     else:
#         cursor = conn.cursor(pymysql.cursors.DictCursor)
#     cursor.execute(sql_text)
#     response = cursor.fetchall()
#     conn.commit()
#     cursor.close()
#     conn.close()
#     return response
#
#
# # 连接总数据库
# def connect_mysql(sql_text, dbs='operation', type='tuple'):
#     conn = pymysql.Connect(host='gz-cdb-lwqgjirt.sql.tencentcdb.com', port=59656, user='beyoungsql', passwd='Bymy2021_',
#                            db=dbs)
#     if type == 'tuple':
#         cursor = conn.cursor()
#     else:
#         cursor = conn.cursor(pymysql.cursors.DictCursor)
#     cursor.execute(sql_text)
#     response = cursor.fetchall()
#     conn.commit()
#     cursor.close()
#     conn.close()
#     return response


country_dict = {"JP": "日本", "AU": "澳洲", "US": "美国", "ES": "西班牙",
                "UK": "英国", "DE": "德国", "FR": "法国", "IT": "意大利",
                "MX": "墨西哥", "CA": "加拿大", "EU": "欧洲", "SG": "东南亚", "MY": "马来西亚"}
site_dict = {"YY": "胤佑", "ZR": "中睿", "AN": "爱瑙", "JH": "京汇", "DNY": "东南亚"}


# 下载文件接口
def download_file(request):
    print("=========下载文件接口==========")
    file_path = request.GET.get("file_path", "")
    file_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + file_path
    file = open(file_path, 'rb')
    response = FileResponse(file)
    file_name = file_path.split("/")[-1]
    response['Content-Disposition'] = 'attachment;filename="%s"' % file_name
    return response


# --------------------------------------------------割割割割割割割割-VAT报告

# VAT报告-查询
def select_vat_report(request):
    # print("=================VAT报告-查询: ", request.GET)
    # month = request.GET.get("month", "")
    # if not month:
    #     download_path_list = []
    #     file_name_list = []
    #     save_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/static/data/VAT_report/"
    #     # 所有月份列表
    #     all_dir_list = os.listdir(save_path)
    #     if len(all_dir_list) > 0:
    #         for i in range(len(all_dir_list)):
    #             # 该月份下的文件名列表
    #             file_name_small_list = os.listdir(save_path + all_dir_list[i])
    #             file_name_list.append(file_name_small_list)
    #             small_list = []
    #             for item in file_name_small_list:
    #                 download_path = "/static/data/VAT_report/" + all_dir_list[i] + "/" + item
    #                 small_list.append(download_path)
    #             download_path_list.append(small_list)
    #         res = {"code": 200, "month_list": all_dir_list,
    #                "file_name": file_name_list, "download_path": download_path_list}
    #         return JsonResponse(res)
    #     else:
    #         res = {"code": 4041, "msg": "目前还有上传的文件"}
    #         return JsonResponse(res)
    # save_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/static/data/VAT_report/" + month
    # if os.path.exists(save_path):
    #     month_list = [month]
    #     file_name_list = [os.listdir(save_path)]
    #     download_path_list = []
    #     small_list = []
    #     for item in file_name_list[0]:
    #         download_path = "/static/data/VAT_report/" + month + "/" + item
    #         small_list.append(download_path)
    #     download_path_list.append(small_list)
    #     res = {"code": 200, "month_list": month_list, "file_name": file_name_list, "download_path": download_path_list}
    #     return JsonResponse(res)
    # else:
    #     res = {"code": 4041, "msg": "没有符合条件的文件"}
    #     return JsonResponse(res)
    print("=================VAT报告-查询: ", request.GET)
    month = request.GET.get("month", "")
    if not month:
        # mouth_dir_name = ""
        urls = "https://www.beyoung.group/vat_file_select/"
        post_res = requests.get(url=urls)
        post_data = json.loads(post_res.text)
        if not post_data['code']:
            res = {"code": 4041, "msg": "没有符合条件的数据"}
            return JsonResponse(res)
        if post_data['code'] != 200:
            res = {"code": 4041, "msg": "查询失败"}
            return JsonResponse(res)
        else:
            receive_data = post_data['data']
            # receive_data = {"2020_11": [file1,file2], "2020-12": []}
            month_list = []
            ret_files_name_list = []
            for k, v in receive_data.items():
                k_list = k.split("_")
                k = "-".join(k_list)
                month_list.append(k)
                ret_files_name_list.append(v)
            res = {"code": 200, "month_list": month_list, "ret_files_name_list": ret_files_name_list}
            return JsonResponse(res)
    else:
        mouth_dir_name = "_".join(month.split("-"))
        send_data = {"mouth_dir_name": mouth_dir_name}
        urls = "https://www.beyoung.group/vat_file_select/"
        post_res = requests.get(url=urls, data=send_data)
        post_data = json.loads(post_res.text)
        if post_data['code'] != 200:
            res = {"code": 4041, "msg": "查询失败"}
            return JsonResponse(res)
        else:
            receive_data = post_data['data']
            month_list = []
            ret_files_name_list = []
            for k, v in receive_data.items():
                k_list = k.split("_")
                k = "-".join(k_list)
                month_list.append(k)
                ret_files_name_list.append(v)
            res = {"code": 200, "month_list": month_list, "ret_files_name_list": ret_files_name_list}
            return JsonResponse(res)


# VAT报告-上传
def upload_vat_report(request):
    print("=================VAT报告-上传: ", request.FILES)
    file = request.FILES.get("file", "")
    mark = request.POST.get("mark", "")
    month = datetime.datetime.strftime(datetime.datetime.today(), "%Y_%m")
    # save_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/static/data/VAT_report/" + month
    # if not os.path.exists(save_path):
    #     os.mkdir(save_path)
    # save_file_path = save_path + "/" + file.name
    # with open(save_file_path, "wb") as fw:
    #     for file_data in file:
    #         fw.write(file_data)

    if mark:
        pass
    else:
        select_only_sql = "select * from report_path where report_time='%s' and report_type='VAT报告';" \
                          % month
        select_only_res = conf_fun.conf_fun.connect_mysql_operation_operation(sql_text=select_only_sql)
        if len(select_only_res) > 0:
            res = {"code": 4042, "msg": "已有文件，是否覆盖？"}
            return JsonResponse(res)

    urls = "https://www.beyoung.group/file_upload/"
    path = {"path": "operation/VAT_report/" + month}
    files = {'file': file}
    post_res = requests.post(url=urls, files=files, data=path)
    post_data = json.loads(post_res.text)
    if post_data['code'] != 200:
        res = {"code": 4041, "msg": "上传失败"}
        return JsonResponse(res)

    # 插入数据库
    insert_sql = "insert into report_path(report_time,report_type,report_path) " \
                 "values('%s', '%s', '%s');" \
                 % (month, "VAT报告", path['path'] + "/" + file.name)
    print("insert_sql: ", insert_sql)
    conf_fun.connect_mysql_operation(sql_text=insert_sql)

    res = {"code": 200}
    return JsonResponse(res)

# --------------------------------------------------割割割割割割割割-库存报告


# 库存报告-查询
def select_inventory_report(request):
    print("===================库存报告-查询: ", request.GET)
    # platform = request.GET.get("platform", "")
    country = request.GET.get("country", "")
    if country:
        for k, v in country_dict.items():
            print(k, v)
            if v == country:
                country = k
    site = request.GET.get("site", "")
    if site:
        for k1, v1 in site_dict.items():
            if v1 == site:
                site = k1
    select_date = request.GET.get("select_date", "")
    select_month = request.GET.get("select_month", "")
    if select_month:
        select_month_end = select_month + "-32"
    page_type = request.GET.get("page_type", "")

    if page_type == "库存报告":
        select_sku_report_sql = "select * from sku_report where id!=''"
        if country:
            select_sku_report_sql += " and countries='" + country + "'"
        if site:
            select_sku_report_sql += " and company='" + site + "'"
        if select_date:
            select_sku_report_sql += " and times='" + select_date + "'"
        if select_month:
            select_sku_report_sql += " and times>'" + select_month + "'" + " and times<'" + select_month_end + "'"
        select_sku_report_sql += ";"
        print("select_sku_report_sql: ", select_sku_report_sql)
        select_sku_report_res = conf_fun.connect_mysql_re(sql_text=select_sku_report_sql)
        print("select_sku_report_res: ", select_sku_report_res)
        print(len(select_sku_report_res))
        if not select_sku_report_res:
            res = {"code": 4041, "msg": "没有符合条件的数据"}
            return JsonResponse(res)

        # 获取数据库表的字段名
        select_key_sql = "select * from sku_report limit 1;"
        select_key_res = conf_fun.connect_mysql_re(sql_text=select_key_sql, type="dict")
        sku_report_key_list = [x for x in select_key_res[0]]
        print(sku_report_key_list)

        # 将数据遍历进excel表
        now_time = int(time.time())
        file_path = os.path.dirname(os.path.dirname(
            os.path.abspath(__file__))) + "/static/data/inventory_report/" + str(now_time) + ".xlsx"
        print(file_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        # 写入表头
        for k in range(len(sku_report_key_list)):
            ws.cell(row=1, column=k + 1).value = sku_report_key_list[k]

        # 写入表格数据
        for i in range(len(select_sku_report_res)):
            for j in range(len(select_sku_report_res[0])):
                ws.cell(row=i + 2, column=j + 1).value = select_sku_report_res[i][j]

        # 保存表格
        wb.save(filename=file_path)

        download_path = "/static/data/inventory_report/" + str(now_time) + ".xlsx"

        res = {"code": 200, "download_path": download_path}

    return JsonResponse(res)


# 库存报告-获取下拉框数据
def get_select_inventory_report(request):
    print("==============库存报告-获取下拉框数据==============")
    page_type = request.GET.get("page_type", "")
    if page_type == "库存报告":
        country_list = []
        site_list = []

        select_sku_report_sql = "select distinct company,countries from sku_report;"
        select_sku_report_res = conf_fun.connect_mysql_re(sql_text=select_sku_report_sql)
        print(select_sku_report_res)

        for item in select_sku_report_res:
            # try:
            #     country_index = country_list.index(item[1])
            #     try:
            #         site_list[country_index].index(item[0])
            #     except IndexError:
            #         site_list.append([site_dict[item[0]]])
            #     except ValueError:
            #         site_list[country_index].append(site_dict[item[0]])
            # except ValueError:
            #     country_list.append(country_dict[item[1]])
            #     site_list.append([site_dict[item[0]]])
            if country_dict[item[1]] not in country_list:
                country_list.append(country_dict[item[1]])
            if site_dict[item[0]] not in site_list:
                site_list.append(site_dict[item[0]])

        res = {"code": 200, "country_list": country_list, "site_list": site_list}
        return JsonResponse(res)

# --------------------------------------------------割割割割割割割割-交易报告/汇总报告/周期结算报告


# 交易报告/汇总报告/周期结算报告-上传
def upload_many_report(request):
    print("=================交易报告/汇总报告/周期结算报告-上传: ", request.FILES, request.POST)
    platform = request.POST.get("platform", "")
    country = request.POST.get("country", "")
    site = request.POST.get("site", "")
    file = request.FILES.get("file", "")
    page_type = request.POST.get("page_type", "")

    target_month = request.POST.get("target_month", "")
    start_date = request.POST.get("start_date", "")
    end_date = request.POST.get("end_date", "")

    mark = request.POST.get("mark", "")

    month_en = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
            'Sept': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}

    if not platform or not country or not site:
        res = {"code": 4041, "msg": "渠道站点国家未填完！"}
        return JsonResponse(res)

    if not file:
        res = {"code": 4041, "msg": "未选择文件！"}
        return JsonResponse(res)

    if page_type == '交易报告' and '.csv' not in file.name:
        res = {"code": 4041, "msg": "交易报告格式必须为csv！"}
        return JsonResponse(res)

    # 是否有记录校验
    if mark:
        pass
    else:
        if page_type == "交易报告" or page_type == "汇总报告":
            select_only_sql = "select * from report_path where platform='%s' and country='%s' " \
                              "and site='%s' and report_time='%s' and report_type='%s';" \
                              % (platform, country, site, target_month, page_type)
            select_only_res = conf_fun.connect_mysql_operation(sql_text=select_only_sql)
            if len(select_only_res) > 0:
                res = {"code": 4042, "msg": "已有文件，是否覆盖？"}
                return JsonResponse(res)
        else:
            select_only_sql = "select * from report_path where platform='%s' and country='%s' " \
                              "and site='%s' and start_time='%s' and end_time='%s' and report_type='%s';" \
                              % (platform, country, site, start_date, end_date, page_type)
            select_only_res = conf_fun.connect_mysql_operation(sql_text=select_only_sql)
            if len(select_only_res) > 0:
                res = {"code": 4042, "msg": "已有文件，是否覆盖？"}
                return JsonResponse(res)

    if page_type == "交易报告":
        path = '/home/by_operate/static/data/transaction/'
        with open(path + file.name, "wb") as f:
            for line in file:
                f.write(line)

        f = open('path + file.name', 'rb')
        r = f.read()

        f_charInfo = chardet.detect(r)
        f.close()
        if country in ('美国', '加拿大', '墨西哥'):
            df = pd.read_csv(path + file.name, encoding=f_charInfo['encoding'], skiprows=6)
            for i in range(df.shape[0]):
                sql = "insert into sale_report(channe,station,country,dates,settlement_id,type,order_id,sku," \
                      "description,quantity,marketplace,account_type,fulfillment,order_city,order_state,order_postal," \
                      "tax_collection_model,product_sales,product_sales_tax,shipping_credits,shipping_credits_tax," \
                      "gift_wrap_credits,gift_wrap_credits_tax,promotional_rebates,promotional_rebates_tax," \
                      "marketplace_withheld_tax,selling_fees,fba_fees,other_transaction_fees,other,total) values " \
                      "('Amazon','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}',"\
                      "'{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')"

                month = month_en[df.iloc[i, 0].split(' ')[0]]
                day = df.iloc[i, 0].split(',')[0].split(' ')[1]
                if len(day) == 1:
                    day = '0' + day
                year = df.iloc[i, 0].split(' ')[2]
                times = df.iloc[i, 0].split(' ')[3]

                if df.iloc[i, 0].split(' ')[4] == 'AM':
                    if times.split(':')[0] == 12:
                        time = '00:' + times.split(':', 1)[1]
                    else:
                        time = times
                else:
                    if times.split(':')[0] == 12:
                        time = times
                    else:
                        time = str(int(times.split(':')[0]) + 12) + times.split(':', 1)[1]
                dates = year + '-' + month + '-' + day + ' ' + time

                sql = sql.format(site, country, dates, df.iloc[i, 1], df.iloc[i, 2], df.iloc[i, 3], df.iloc[i, 4],
                                 df.iloc[i, 5], df.iloc[i, 6], df.iloc[i, 7], df.iloc[i, 8], df.iloc[i, 9],
                                 df.iloc[i, 10], df.iloc[i, 11], df.iloc[i, 12], df.iloc[i, 13], df.iloc[i, 14],
                                 df.iloc[i, 15], df.iloc[i, 16], df.iloc[i, 17], df.iloc[i, 18], df.iloc[i, 19],
                                 df.iloc[i, 20], df.iloc[i, 21], df.iloc[i, 22], df.iloc[i, 23], df.iloc[i, 24],
                                 df.iloc[i, 25], df.iloc[i, 26], df.iloc[i, 27])
                conf_fun.connect_mysql_operation(sql, dbs='financial')
        elif country in ('英国', '德国', '法国', '意大利', '西班牙'):
            df = pd.read_csv(path + file.name, encoding=f_charInfo['encoding'], skiprows=6)
            for i in range(df.shape[0]):
                sql = "insert into sale_report(channe,station,country,dates,settlement_id,type,order_id,sku," \
                      "description,quantity,marketplace,fulfillment,order_city,order_state,order_postal," \
                      "tax_collection_model,product_sales,product_sales_tax,shipping_credits,shipping_credits_tax," \
                      "gift_wrap_credits,gift_wrap_credits_tax,promotional_rebates,promotional_rebates_tax," \
                      "marketplace_withheld_tax,selling_fees,fba_fees,other_transaction_fees,other,total) values " \
                      "('Amazon','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}'," \
                      "'{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')"

                month = month_en[df.iloc[i, 0].split(' ')[1]]
                day = df.iloc[i, 0].split(' ')[0]
                if len(day) == 1:
                    day = '0' + day
                year = df.iloc[i, 0].split(' ')[2]
                times = df.iloc[i, 0].split(' ')[3]

                dates = year + '-' + month + '-' + day + ' ' + times

                sql = sql.format(site, country, dates, df.iloc[i, 1], df.iloc[i, 2], df.iloc[i, 3], df.iloc[i, 4],
                                 df.iloc[i, 5], df.iloc[i, 6], df.iloc[i, 7], df.iloc[i, 8], df.iloc[i, 9],
                                 df.iloc[i, 10], df.iloc[i, 11], df.iloc[i, 12], df.iloc[i, 13], df.iloc[i, 14],
                                 df.iloc[i, 15], df.iloc[i, 16], df.iloc[i, 17], df.iloc[i, 18], df.iloc[i, 19],
                                 df.iloc[i, 20], df.iloc[i, 21], df.iloc[i, 22], df.iloc[i, 23], df.iloc[i, 24],
                                 df.iloc[i, 25], df.iloc[i, 26])
                conf_fun.connect_mysql_operation(sql, dbs='financial')

        elif country == '日本':

            df = pd.read_csv(path + file.name, encoding=f_charInfo['encoding'], skiprows=6)
            for i in range(df.shape[0]):
                sql = "insert into sale_report(channe,station,country,dates,settlement_id,type,order_id,sku," \
                      "description,quantity,marketplace,account_type,fulfillment,order_city,order_state,order_postal," \
                      "tax_collection_model,product_sales,product_sales_tax,shipping_credits,shipping_credits_tax," \
                      "gift_wrap_credits,gift_wrap_credits_tax,promotional_rebates,promotional_rebates_tax," \
                      "marketplace_withheld_tax,selling_fees,fba_fees,other_transaction_fees,other,total) values " \
                      "('Amazon','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}',"\
                      "'{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')"
                dates = df.iloc[i, 0].replace('/', '-').rsplit(' ', 1)[0]

                sql = sql.format(site, country, dates, df.iloc[i, 1], df.iloc[i, 2], df.iloc[i, 3], df.iloc[i, 4],
                                 df.iloc[i, 5], df.iloc[i, 6], df.iloc[i, 7], df.iloc[i, 8], df.iloc[i, 9],
                                 df.iloc[i, 10], df.iloc[i, 11], df.iloc[i, 12], df.iloc[i, 13], df.iloc[i, 14],
                                 df.iloc[i, 15], df.iloc[i, 16], df.iloc[i, 17], df.iloc[i, 18], df.iloc[i, 19],
                                 df.iloc[i, 20], df.iloc[i, 21], df.iloc[i, 22], df.iloc[i, 23], df.iloc[i, 24],
                                 df.iloc[i, 25], df.iloc[i, 26], df.iloc[i, 27])
                conf_fun.connect_mysql_operation(sql, dbs='financial')
        # 澳洲
        else:
            df = pd.read_csv(path + file.name, encoding=f_charInfo['encoding'], skiprows=6)
            for i in range(df.shape[0]):
                sql = "insert into sale_report(channe,station,country,dates,settlement_id,type,order_id,sku," \
                      "description,quantity,marketplace,fulfillment,order_city,order_state,order_postal," \
                      "product_sales,shipping_credits,gift_wrap_credits,promotional_rebates,sales_tax_collected," \
                      "low_value_goods,selling_fees,fba_fees,other_transaction_fees,other,total) values " \
                      "('Amazon','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}',"\
                      "'{}','{}','{}','{}','{}','{}','{}','{}')"

                month = month_en[df.iloc[i, 0].split(' ')[1]]
                day = df.iloc[i, 0].split(' ')[0]
                if len(day) == 1:
                    day = '0' + day
                year = df.iloc[i, 0].split(' ')[2]
                times = df.iloc[i, 0].split(' ')[3]

                if df.iloc[i, 0].split(' ')[4] == 'am':
                    if times.split(':')[0] == 12:
                        time = '00:' + times.split(':', 1)[1]
                    else:
                        time = times
                else:
                    if times.split(':')[0] == 12:
                        time = times
                    else:
                        time = str(int(times.split(':')[0]) + 12) + times.split(':', 1)[1]
                dates = year + '-' + month + '-' + day + ' ' + time

                sql = sql.format(site, country, dates, df.iloc[i, 1], df.iloc[i, 2], df.iloc[i, 3], df.iloc[i, 4],
                                 df.iloc[i, 5], df.iloc[i, 6], df.iloc[i, 7], df.iloc[i, 8], df.iloc[i, 9],
                                 df.iloc[i, 10], df.iloc[i, 11], df.iloc[i, 12], df.iloc[i, 13], df.iloc[i, 14],
                                 df.iloc[i, 15], df.iloc[i, 16], df.iloc[i, 17], df.iloc[i, 18], df.iloc[i, 19],
                                 df.iloc[i, 20], df.iloc[i, 21], df.iloc[i, 22], df.iloc[i, 23])
                conf_fun.connect_mysql_operation(sql, dbs='financial')



        # save_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/static/data/transaction_report/"
        # save_path += platform + country + site + "/"
        # if not os.path.exists(save_path):
        #     os.mkdir(save_path)
        # save_path += now_month + "/"
        # if not os.path.exists(save_path):
        #     os.mkdir(save_path)
        # save_path += file.name
        #
        # with open(save_path, "wb") as fw:
        #     for file_data in file:
        #         fw.write(file_data)
        #
        # static_path = "/static/data/transaction_report/" + platform + country + site + "/" + now_month
        # static_path += "/" + file.name

        # 上传文件到静态文件服务器
        urls = "https://www.beyoung.group/file_upload/"
        path = {"path": "operation/transaction_report/" + platform + "/" + site + "/" + country + "/"}
        files = {'file': file}
        post_res = requests.post(url=urls, files=files, data=path)
        post_data = json.loads(post_res.text)
        if post_data['code'] != 200:
            res = {"code": 4041, "msg": "上传失败"}
            return JsonResponse(res)

        # 插入数据库
        insert_sql = "insert into report_path(platform,country,site,report_time,report_type,report_path) " \
                     "values('%s', '%s', '%s', '%s', '%s', '%s');" \
                     % (platform, country, site, target_month, "交易报告", path['path'] + file.name)
        print("insert_sql: ", insert_sql)
        conf_fun.connect_mysql_operation(sql_text=insert_sql)

    if page_type == "汇总报告":
        # save_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/static/data/gather_report/"
        # save_path += platform + country + site + "/"
        # if not os.path.exists(save_path):
        #     os.mkdir(save_path)
        # save_path += now_month + "/"
        # if not os.path.exists(save_path):
        #     os.mkdir(save_path)
        # save_path += file.name
        #
        # with open(save_path, "wb") as fw:
        #     for file_data in file:
        #         fw.write(file_data)

        # static_path = "/static/data/gather_report/" + platform + country + site + "/" + now_month
        # static_path += "/" + file.name

        # 插入数据库
        # insert_sql = "insert into report_path(platform,country,site,report_time,report_type,report_path) " \
        #              "values('%s', '%s', '%s', '%s', '%s', '%s');" \
        #              % (platform, country, site, now_month, "汇总报告", static_path)
        # print("insert_sql: ", insert_sql)
        # conf_fun.connect_mysql_operation(sql_text=insert_sql)
        urls = "https://www.beyoung.group/file_upload/"
        path = {"path": "operation/gather_report/" + platform + "/" + site + "/" + country + "/"}
        files = {'file': file}
        post_res = requests.post(url=urls, files=files, data=path)
        post_data = json.loads(post_res.text)
        if post_data['code'] != 200:
            res = {"code": 4041, "msg": "上传失败"}
            return JsonResponse(res)

        # 插入数据库
        insert_sql = "insert into report_path(platform,country,site,report_time,report_type,report_path) " \
                     "values('%s', '%s', '%s', '%s', '%s', '%s');" \
                     % (platform, country, site, target_month, "汇总报告", path['path'] + file.name)
        print("insert_sql: ", insert_sql)
        conf_fun.connect_mysql_operation(sql_text=insert_sql)

    if page_type == "结算周期报告":
        if not start_date or not end_date:
            res = {"code": 4041, "msg": "时间未填完整!"}
            return JsonResponse(res)
        # save_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/static/data/cycle_report/"
        # save_path += platform + country + site + "/"
        # if not os.path.exists(save_path):
        #     os.mkdir(save_path)
        # save_path += start_date + "_" + end_date + "/"
        # if not os.path.exists(save_path):
        #     os.mkdir(save_path)
        # save_path += file.name
        #
        # with open(save_path, "wb") as fw:
        #     for file_data in file:
        #         fw.write(file_data)
        #
        # static_path = "/static/data/cycle_report/" + platform + country + site + "/" + start_date + "_" + end_date
        # static_path += "/" + file.name
        #
        # # 插入数据库
        # insert_sql = "insert into report_path(platform,country,site,start_time,end_time,report_type,report_path) " \
        #              "values('%s', '%s', '%s', '%s', '%s', '%s', '%s');" \
        #              % (platform, country, site, start_date, end_date, "结算周期报告", static_path)
        # print("insert_sql: ", insert_sql)
        # conf_fun.connect_mysql_operation(sql_text=insert_sql)

        urls = "https://www.beyoung.group/file_upload/"
        path = {"path": "operation/cycle_report/" + platform + "/" + site + "/" + country + "/"}
        files = {'file': file}
        post_res = requests.post(url=urls, files=files, data=path)
        post_data = json.loads(post_res.text)
        if post_data['code'] != 200:
            res = {"code": 4041, "msg": "上传失败"}
            return JsonResponse(res)

        # 插入数据库
        insert_sql = "insert into report_path(platform,country,site,start_time,end_time,report_type,report_path) " \
                     "values('%s', '%s', '%s', '%s', '%s', '%s', '%s');" \
                     % (platform, country, site, start_date, end_date, "结算周期报告", path['path'] + file.name)
        print("insert_sql: ", insert_sql)
        conf_fun.connect_mysql_operation(sql_text=insert_sql)

    res = {"code": 200}
    return JsonResponse(res)


# 交易报告/汇总报告/结算周期报告-查询
def select_many_report(request):
    print("=================交易报告/汇总报告/结算周期报告-查询: ", request.GET)
    platform = request.GET.get("platform", "")
    country = request.GET.get("country", "")
    site = request.GET.get("site", "")
    select_month = request.GET.get("select_month", "")
    start_time = request.GET.get("start_date", "")
    end_time = request.GET.get("end_date", "")
    page_type = request.GET.get("page_type", "")

    if page_type == "结算周期报告":
        select_report_path_sql = "select * from report_path " \
                                 "where platform='%s' and country='%s' and site='%s' and report_type='%s'" \
                                 % (platform, country, site, "结算周期报告")
        if start_time:
            select_report_path_sql += " and start_time>='" + start_time + "'"
        if end_time:
            select_report_path_sql += " and end_time<='" + end_time + "'"
    else:
        select_report_path_sql = "select * from report_path where platform='%s' and country='%s' and site='%s'" \
                                 % (platform, country, site)
        if page_type == "交易报告":
            select_report_path_sql += " and report_type='" + "交易报告" + "'"
        if page_type == "汇总报告":
            select_report_path_sql += " and report_type='" + "汇总报告" + "'"
        if select_month:
            select_report_path_sql += " and report_time='" + select_month + "'"
    select_report_path_sql += ";"
    print("select_report_path_sql: ", select_report_path_sql)
    select_report_path_res = conf_fun.connect_mysql_operation(sql_text=select_report_path_sql)
    print("select_report_path_res: ", select_report_path_res)
    if len(select_report_path_res) > 0:
        res = {"code": 200, "data": select_report_path_res}
        return JsonResponse(res)
    else:
        res = {"code": 4041, "msg": "没有符合条件的数据!"}
        return JsonResponse(res)


# 交易报告/汇总报告/结算周期报告-获取下拉框数据
def get_select_many_report(request):
    print("========交易报告/汇总报告/结算周期报告-获取下拉框数据=========")
    select_sql = "select distinct platform,country,site from report_path;"
    select_res = conf_fun.connect_mysql_operation(sql_text=select_sql)

    platform_list = []
    country_list = []
    site_list = []

    for item in select_res:
        if item[0] not in platform_list:
            platform_list.append(item[0])
        if item[1] not in platform_list:
            country_list.append(item[1])
        if item[2] not in platform_list:
            site_list.append(item[2])

    res = {"code": 200, "platform_list": platform_list, "country_list": country_list, "site_list": site_list}
    return JsonResponse(res)


# 商品核对时间 商品信息
def commodity_check_time(request):
    print("\n", "修改商品核对时间", "\n")
    data_id = request.GET.get("data_id", "")
    now_time = datetime.datetime.strftime(datetime.datetime.now() + datetime.timedelta(hours=8), "%Y-%m-%dT%H:%M:%S")
    now_time_str = now_time.replace("T", "\n")
    update_sql = "update commodity_information set check_time='%s' where id='%s';" % (now_time_str, data_id)
    conf_fun.connect_mysql_operation(sql_text=update_sql, dbs='operation')

    res = {"code": 200}
    return JsonResponse(res)
























