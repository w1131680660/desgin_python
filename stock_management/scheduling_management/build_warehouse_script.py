import os
import re
import time
from itertools import groupby
from operator import itemgetter

import pandas as pd
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from stock_management import scheduling_settings
from stock_management.scheduling_management import post_generation
from stock_management.scheduling_management import scheduling_func

from settings import conf_fun

site_dict = {
    '胤佑': 'YY',
    '京汇': 'JH',
    '中睿': 'ZR',
    '爱瑙': 'AN',
    '利百锐': 'LBR',
    '笔漾教育': 'BYJY',
    '景瑞': 'JR'}

country_dict = {'澳洲': 'AU',
                '加拿大': 'CA',
                '西班牙': 'ES',
                '英国': 'UK',
                '法国': 'FR',
                '意大利': 'IT',
                '德国': 'DE',
                '日本': 'JP',
                '美国': 'US',
                '墨西哥': 'MX'
                }


def create_file(file_path, container_code, country, site):
    if country in ['英国', '法国', '德国', '意大利', '西班牙']:
        country_re = '欧洲'
    else:
        country_re = country
    localtime = time.strftime("%Y-%m-%d", time.localtime())
    if file_path:
        pd_data = pd.read_csv(file_path, sep='\t', error_bad_lines=False, header=0)

        fba_num = pd_data.columns.values[1]
        shipment_name = pd_data.columns.values[0]
        shipment_code = pd_data.columns.values[1]

        Shipment_information = "货件名称/编号 {0}:{1} \n".format(shipment_name, shipment_code)
        for k, v in pd_data.iterrows():
            if k < 3:
                Shipment_information += "{0}:{1} \n".format(v[0], v[1])
    else:
        Shipment_information =''
        file_name, fba_num ='',''
    sql = ''' SELECT * FROM order_distribution WHERE country='{0}' and store ='{1}' and 
              container_code ='{2}' and `status` ='0' '''.format(country, site, container_code)

    # re_data = conf_fun.connect_mysql_operation(sql, dbs='supply_chain', type='dict')

    con_sql = '''  SELECT DISTINCT sc.country,sc.warehouse_type,sc.container_type,sc.site,op.calculation_file_path 
    FROM schedule_container as sc,operating_data as op 
    WHERE sc.container_num = op.container_num and sc.container_num ='{0}' '''.format(container_code)
    print('这是89行\n\n', con_sql)

    con_data = conf_fun.connect_mysql_operation(con_sql, type='dict')
    con_data = con_data[0]

    x = con_data.get('calculation_file_path').split('/')[-1].split('-')[1]
    name = con_data.get('calculation_file_path').split('/')[-1].split('-')[2]
    container_type = con_data.get('container_type')
    warehouse_type = con_data.get('warehouse_type')

    country_low = country_dict.get(country)
    site_low = site_dict.get(site)
    file_name = ''' {0}-{1}-{3}{4}{2}'''.format(container_code, x, container_type, site_low, country_low)

    sql = ''' select DISTINCT pp.*,co.fnsku from (SELECT ord.sku,ord.product_name, ord.distribution_num AS num, pro.product_code,ord.factory,pro.volume,pro.product_package_size,pro.product_weight
            FROM order_distribution AS ord INNER JOIN operation.product_message AS pro ON ord.product_number = pro.product_code
            AND ord.container_code = '{0}' AND ord.country = '{1}' AND ord.store = '{2}' ) as pp INNER JOIN operation.commodity_information  as co 
            on pp.sku = co.sku and co.country = '{3}' and co.site ='{2}'
             ORDER BY pp.factory DESC,pp.product_name'''.format(container_code, country, site, country_re)
    # FNSKU 是产品条码
    print(109, 'sql\n', sql)
    sql_data = conf_fun.connect_mysql_operation(sql, dbs='supply_chain', type='dict')
    re_list = []
    pattern = '\d+'
    for i in sql_data:
        product_name = i.get('product_name')
        if product_name:
            re_name = re.findall(pattern, product_name)
            re_name = int(re_name[0]) if re_name else 0
            i['re_name'] = re_name
        product_number = i.get('product_code')
        if product_number:
            product_number_list = product_number.split('-')[0:-1]
            product_number_re = '-'.join(product_number_list)
            i['product_number_re'] = product_number_re
        re_list.append(i)
    re_list = sorted(re_list, key=itemgetter('factory', 'product_number_re', 're_name'))
    sql_data = re_list
    data_dict = {}
    for factory, items in groupby(sql_data, key=itemgetter('factory')):
        if factory not in data_dict.keys():
            data_dict[factory] = []
        for i in items:
            factory_name = i.get('factory')
            sku = i.get('sku')
            product_name = i.get('product_name')
            num = i.get('num')
            product_package_size = i.get('product_package_size')  # 尺寸
            try:
                volume = round(float(i.get('volume')), 2) if i.get('volume') else ''  # 体积
            except:
                volume = ''
            product_weight = i.get('product_weight')  # 重量
            fnsku = i.get('fnsku')  # fnsku
            if fba_num :
                Follow = " 编号 : {0} \n 箱贴 : {1}".format(file_name, fba_num)
            else: Follow =''
            dict_1 = {'sku': sku, 'product_name': product_name, 'num': num,
                      'volume': volume, 'product_weight': product_weight, 'fnsku': fnsku,
                      'follow': Follow, 'product_package_size': product_package_size}
            data_dict[factory].append(dict_1)

    for k, v in data_dict.items():
        print(129, '----------', len(v))
        wb = Workbook()
        ws = wb.active
        ws.title = '建仓计划表'

        ws.merge_cells('A1:K1')  # 这个对指定范围的单元格进行合并
        ws['A1'] = ''' {0}-{1}-{3}{4}{2}'''.format(container_code, x, container_type, site_low, country_low)
        ws.merge_cells('A2:B2')
        ws['A2'] = '货物编码'
        ws.merge_cells('C2:E2')
        ws['C2'] = file_name
        ws['F2'] = '店名'

        ws['A3'], ws['B3'], ws['C3'], ws['D3'], ws['E3'], ws['F3'], ws['G3'], ws['H3'], ws['I3'], ws['J3'], ws['K3'] = \
            '序号', 'sku', '货名', '数量', '尺寸', '体积(m^3)', '重量(KG)', '产品条码', '箱贴', '货件信息', '备注'

        font = Font(name='宋体', size=11)
        font_1 = Font(name='宋体', size=15, bold=True)
        alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        fill = PatternFill(fill_type="solid", fgColor="A6A6A6")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ws.cell(row=1, column=1).font = font_1
        ws.cell(row=1, column=1).alignment = alignment
        ws.cell(row=1, column=1).fill = fill

        ws.cell(row=2, column=3).font = font_1

        ws.cell(row=2, column=4).font = font_1

        for i in range(1, 7):
            ws.cell(row=2, column=i).font = font_1
            ws.cell(row=2, column=i).alignment = alignment
            if i in [1, 2, 3, 4]:
                ws.cell(row=2, column=4).border = thin_border
            if i != 3:
                ws.cell(row=2, column=i).fill = fill

        for i in range(1, 12):
            ws.cell(row=3, column=i).alignment = alignment
            ws.cell(row=3, column=i).fill = fill
            ws.cell(row=3, column=i).font = font_1
            ws.cell(row=3, column=i).border = thin_border
            ws.cell(row=1, column=i).border = thin_border

        for i in range(1, 4):
            ws.cell(row=2, column=i).border = thin_border

        count_num = 0  # 求和数量
        count_volume = 0  # 求和体积
        count_weight = 0  # 求和体积
        for index, dict_2 in enumerate(v, 1):
            ws["A%s" % (index + 3)] = index

            ws['B%s' % (index + 3)] = dict_2.get('sku')

            ws['C%s' % (index + 3)], ws['D%s' % (index + 3)] = \
                dict_2.get('product_name'), dict_2.get('num')
            ws['E%s' % (index + 3)], ws['F%s' % (index + 3)], ws['G%s' % (index + 3)] = \
                dict_2.get('product_package_size'), dict_2.get('volume'), dict_2.get('product_weight')

            ws['H%s' % (index + 3)], ws['I%s' % (index + 3)], ws['J%s' % (index + 3)], ws['K%s' % (index + 3)] = \
                dict_2.get('fnsku'), dict_2.get('follow'), '', ''

            count_num += float(dict_2.get('num')) if dict_2.get('num') else 0
            count_volume += float(dict_2.get('product_volume')) if dict_2.get('product_volume') else 0
            count_weight += float(dict_2.get('product_weight')) if dict_2.get('product_weight') else 0
            for q in range(1, 12):
                if q not in [9, 10]:
                    ws.cell(row=index + 3, column=q).alignment = alignment
                else:
                    ws.cell(row=index + 3, column=q).alignment = left_alignment
                ws.cell(row=index + 3, column=q).border = thin_border
        ws["A%s" % (4 + len(v))] = '总计'

        ws["D%s" % (4 + len(v))] = count_num
        ws["F%s" % (4 + len(v))] = count_volume
        ws["G%s" % (4 + len(v))] = count_weight
        for i in range(1, 12):
            if q not in [9, 10]:
                ws.cell(row=(4 + len(v)), column=i).alignment = alignment
            else:
                ws.cell(row=(4 + len(v)), column=q).alignment = left_alignment
            ws.cell(row=(4 + len(v)), column=i).border = thin_border
        for i in range(1, 4 + len(v)):
            if i >= 4:

                if len(v) == 1:
                    print(i)
                    ws.row_dimensions[i].height = 120
                else:
                    ws.row_dimensions[i].height = 55 if 1 < len(v) <= 2 else 40
            else:
                ws.row_dimensions[i].height = 40
        ws.merge_cells('J4:J%s' % (3 + len(v)))
        ws['J4'] = Shipment_information
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 40

        date = time.strftime('%m%d', time.localtime())
        filename_1 = '''{0}{1}-{2}-{3}-{4}-{5}-{6}-{7}-{8}-建仓信息确认表.xlsx''' \
            .format(site, country, x, container_code, warehouse_type, container_type, name, k, date)
        print(filename_1)
        dirs = "static/data/operating_data/build_warehouse/"
        if not os.path.exists(dirs):
            os.makedirs(dirs)
        wb.save(dirs + filename_1)

        # 这里还是将建仓信息表发给老系统同时生成条码附件和条码

        establish__str = ''
        update_str = ''
        data = {'container_num': container_code, 'site': site, 'country': country,
                'file_name': open(dirs + filename_1, 'rb')}
        scheduling_func.master_upload_file(open(dirs + filename_1, 'rb'), 'establish_warehouse')
        sql_value = ['establish_warehouse', 'establish_warehouse_date', 'establish_warehouse',
                     'establish_warehouse_state']

        path3 = os.path.join(r'operation/operating_data/', '%s/%s' % (sql_value[0], filename_1.strip()))
        sql = ''' UPDATE operating_data SET {0} {1} '''
        where_str = ''' WHERE container_num = '{0}' and  site = '{1}'  and   country = '{2}' and factory = '{3}' ''' \
            .format(container_code, site, country, k)
        establish__str += " {0} = '{1}',{2}= '{3}', {4} ='{5}' " \
            .format(sql_value[1], localtime, sql_value[2], path3, sql_value[3], '未审核')

        bar_code_path, bar_code_path_new, error_msg = post_generation.create_post_generation(data)

        date_time_1 = localtime
        date_time = localtime
        state_type = '已审核'
        state_type_1 = '未审核'
        if not bar_code_path_new:
            bar_code_path_new = ''
            date_time = ''
            state_type = ''
        if not bar_code_path:
            bar_code_path = ''
            date_time_1 = ''
            state_type_1 = ''

        update_str += "{0} = '{1}', {2} ='{3}',{4} ='{5}',{6} = '{7}', {8} ='{9}',{10} ='{11}' " \
            .format('bar_code', bar_code_path, 'bar_code_date', date_time_1, 'bar_code_state', state_type_1,
                    'bar_code_new', bar_code_path_new, 'bar_code_date_new', date_time,
                    'bar_code_state_new', state_type)
        establish_sql = sql.format(establish__str, where_str)
        bar_sql = sql.format(update_str, where_str)
        print('\n\n建仓信息表的sql', establish_sql)
        print('条码的sql', bar_sql)

        conf_fun.connect_mysql_operation(establish_sql)
        conf_fun.connect_mysql_operation(bar_sql)

        # print(bar_code_path, bar_code_path_new, error_msg )




from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
def warehouse_script(request):
    data = request.POST
    files = request.FILES.get('files')
    container_code = data.get('container_num')
    country = data.get('country')
    site = data.get('site')
    ret = {'code': 200, 'msg': '无'}
    file_name = files.name
    if 'tsv' not in file_name:
        ret['code'] = 500
        ret['error_msg'] = '请上传正确的文件，文件格式为tsv'
        print(file_name)
    else:
        try:
            create_file(files, container_code, country, site)
        except:
            ret['code'] = 500
            ret['error_msg'] = '建仓失败，请查看对应的tsv文件内容以及订单分配情况是否正确'
    print(data, files)

    return JsonResponse(ret)

# 生成海外仓库
@csrf_exempt
def fbm_warehouse(request):
    data = request.GET
    # files = request.FILES.get('files')
    container_code = data.get('container_num')
    country = data.get('country')
    site = data.get('site')

    create_file('', container_code, country, site)
    ret ={'code':200}

    return JsonResponse(ret)

def get_distribution_sku(request):
    data = request.GET
    container_code = data.get('container_num')
    country = data.get('country')
    site = data.get('site')
    factory = data.get('factory')


    sql = ''' SELECT distribution_num as plan_order,sku   FROM order_distribution WHERE container_code ='{0}' 
              AND factory ='{1}' and country ='{2}' and store ='{3}' ''' \
        .format(container_code, factory, country, site)

    re_data = conf_fun.connect_mysql_supply(sql, type='dict')
    ret = {'code': 200, 're_sku_data': re_data}
    return JsonResponse(ret)

